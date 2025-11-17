#!/usr/bin/env python3
"""
Простой конвертер CSV в Excel без pandas
"""

def csv_to_excel_simple():
    """Конвертирует CSV в Excel используя базовые библиотеки"""
    
    csv_path = r'c:\Projects\Anonymizer\unified_document_service\patterns\sensitive_patterns_full.csv'
    excel_path = r'c:\Projects\Anonymizer\unified_document_service\patterns\sensitive_patterns.xlsx'
    
    print(f"Конвертируем {csv_path} в {excel_path}")
    
    try:
        # Читаем CSV файл
        with open(csv_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        print(f"Прочитано строк: {len(lines)}")
        
        # Создаем простой Excel файл
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Patterns"
            
            for row_idx, line in enumerate(lines, 1):
                # Простое разделение по запятым
                parts = line.strip().split(',')
                if len(parts) >= 4:
                    category = parts[0]
                    pattern = parts[1] 
                    # Собираем описание (может содержать запятые)
                    description = ','.join(parts[2:-1])
                    confidence = parts[-1]
                    
                    ws.cell(row=row_idx, column=1, value=category)
                    ws.cell(row=row_idx, column=2, value=pattern)
                    ws.cell(row=row_idx, column=3, value=description)
                    ws.cell(row=row_idx, column=4, value=float(confidence) if confidence.replace('.','').isdigit() else confidence)
            
            wb.save(excel_path)
            print(f"✅ Excel файл создан: {excel_path}")
            return True
            
        except ImportError:
            print("❌ openpyxl не доступен. Создаем Excel через csv модуль...")
            
            # Альтернативный способ - создание через pandas если доступен
            try:
                import pandas as pd
                df = pd.read_csv(csv_path)
                df.to_excel(excel_path, index=False)
                print(f"✅ Excel файл создан через pandas: {excel_path}")
                return True
                
            except ImportError:
                print("❌ pandas также недоступен")
                return False
                
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return False

if __name__ == "__main__":
    csv_to_excel_simple()